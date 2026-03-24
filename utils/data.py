"""
data.py — All read/write operations to tracker.xlsx
Each sheet maps to one domain: Savings, Viewings, UpcomingViewings
"""
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA_DIR = Path(__file__).parent.parent / "data"
EXCEL_PATH = DATA_DIR / "tracker.xlsx"

SHEETS = {
    "Savings":          ["key", "value"],
    "Viewings":         ["id", "address", "date", "area", "listed_price", "size_sqm",
                         "avgift", "outcome", "num_bid_rounds", "final_price", "rating", "notes",
                         "hemnet_url", "booli_url"],
    "UpcomingViewings": ["id", "address", "datetime", "area", "asking_price", "notes"],
}

DEFAULT_SETTINGS = {
    "p1_name": "",
    "p2_name": "",
    "p1_current": 0,
    "p2_current": 0,
    "p1_monthly": 0,
    "p2_monthly": 0,
    "loan_amount": 0,
    "loan_expiry": "",
    "loan_bank": "",
    "apartment_price": 0,
    "down_pct": 10,
}

# ── Styling helpers ──────────────────────────────────────────────────────────
def _hdr_fill():
    return PatternFill("solid", fgColor="2E7D5E")

def _style_header(ws, row, ncols):
    thin = Side(style="thin", color="D1D5DB")
    b = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = _hdr_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = b

def _init_workbook():
    """Create a fresh tracker.xlsx with all sheets and headers."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, cols in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        for c, col in enumerate(cols, 1):
            ws.cell(row=1, column=c, value=col)
        _style_header(ws, 1, len(cols))
        ws.freeze_panes = "A2"

    # Pre-populate default Settings
    ws_sav = wb["Savings"]
    for r, (k, v) in enumerate(DEFAULT_SETTINGS.items(), 2):
        ws_sav.cell(row=r, column=1, value=k)
        ws_sav.cell(row=r, column=2, value=v)

    wb.save(EXCEL_PATH)

def _ensure_excel():
    if not EXCEL_PATH.exists():
        _init_workbook()

# ── Generic read/write ────────────────────────────────────────────────────────
def read_sheet(sheet_name: str) -> pd.DataFrame:
    _ensure_excel()
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name=sheet_name, engine="openpyxl")
        return df
    except Exception:
        return pd.DataFrame(columns=SHEETS[sheet_name])

def _write_df_to_sheet(df: pd.DataFrame, sheet_name: str):
    """Overwrite a single sheet while preserving all other sheets."""
    _ensure_excel()
    wb = load_workbook(EXCEL_PATH)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    cols = list(df.columns)
    for c, col in enumerate(cols, 1):
        ws.cell(row=1, column=c, value=col)
    _style_header(ws, 1, len(cols))
    ws.freeze_panes = "A2"
    for r, row in enumerate(df.itertuples(index=False), 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    # Restore sheet order
    desired_order = list(SHEETS.keys())
    current = wb.sheetnames
    ordered = [s for s in desired_order if s in current] + [s for s in current if s not in desired_order]
    wb._sheets = [wb[s] for s in ordered]
    wb.save(EXCEL_PATH)

# ── Settings ──────────────────────────────────────────────────────────────────
def load_settings() -> dict:
    df = read_sheet("Savings")
    kv = df[df["key"].notna()].set_index("key")["value"].to_dict() if not df.empty else {}
    settings = dict(DEFAULT_SETTINGS)
    for k in DEFAULT_SETTINGS:
        if k in kv and pd.notna(kv[k]):
            settings[k] = kv[k]
    # Type coercions
    for int_key in ["p1_current","p2_current","p1_monthly","p2_monthly","loan_amount","apartment_price"]:
        try: settings[int_key] = int(float(settings[int_key]))
        except: settings[int_key] = 0
    try: settings["down_pct"] = float(settings["down_pct"])
    except: settings["down_pct"] = 15.0
    return settings

def save_settings(settings: dict):
    rows = [{"key": k, "value": v} for k, v in settings.items()]
    df = pd.DataFrame(rows, columns=["key", "value"])
    _write_df_to_sheet(df, "Savings")

# ── Viewings ──────────────────────────────────────────────────────────────────
def load_viewings() -> pd.DataFrame:
    df = read_sheet("Viewings")
    if df.empty:
        return pd.DataFrame(columns=SHEETS["Viewings"])
    return df.fillna("")

def save_viewing(row: dict):
    df = load_viewings()
    new_row = pd.DataFrame([row])
    df = pd.concat([df, new_row], ignore_index=True)
    _write_df_to_sheet(df, "Viewings")

# ── Upcoming Viewings ─────────────────────────────────────────────────────────
def load_upcoming() -> pd.DataFrame:
    df = read_sheet("UpcomingViewings")
    return df.fillna("") if not df.empty else pd.DataFrame(columns=SHEETS["UpcomingViewings"])

def save_upcoming(row: dict):
    df = load_upcoming()
    new_row = pd.DataFrame([row])
    df = pd.concat([df, new_row], ignore_index=True)
    _write_df_to_sheet(df, "UpcomingViewings")

def delete_upcoming(uv_id):
    df = load_upcoming()
    df = df[df["id"] != uv_id]
    _write_df_to_sheet(df, "UpcomingViewings")

